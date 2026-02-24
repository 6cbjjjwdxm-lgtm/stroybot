import os
import asyncio
import logging
import base64
import tempfile
import shutil
import json
import pytz

from apscheduler.triggers.cron import CronTrigger
from datetime import datetime

from openpyxl import Workbook, load_workbook

from dotenv import load_dotenv
from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
    KeyboardButton,
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
)
from openai import OpenAI

import rag_engine


# -------------------- ENV / PATHS --------------------
load_dotenv()

DATA_DIR = os.getenv("DATA_DIR", "/var/data")
DATA_DIR = os.path.abspath(DATA_DIR)
os.makedirs(DATA_DIR, exist_ok=True)

DEADLINES_FILE = os.path.join(DATA_DIR, "deadlines.json")
PROGRESS_STATE_FILE = os.path.join(DATA_DIR, "progress_state.json")

PROJECTS_DIR = os.path.join(DATA_DIR, "StroyBot_Files")
os.makedirs(PROJECTS_DIR, exist_ok=True)

# -------------------- LOGGING --------------------
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
# гасим INFO-логи httpx, чтобы не светить токен в URL
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("openai").setLevel(logging.WARNING)

logger = logging.getLogger(__name__)

# -------------------- SECRETS --------------------
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

if not TELEGRAM_TOKEN:
    raise RuntimeError("TELEGRAM_TOKEN is not set")
if not OPENAI_API_KEY:
    raise RuntimeError("OPENAI_API_KEY is not set")

client = OpenAI(api_key=OPENAI_API_KEY)

# -------------------- CONFIG --------------------
ADMIN_USER_IDS = {459980503, 5130953211, 1229215603}

GROUPS_CONFIG = {
    "Мосрентген 28": {
        "chat_id": -5207136504,
        "systems": ["Ремонт подвала", "Ремонт крыши", "Ремонт ЭОМ", "Замена мусоропровода"],
    },
    "Мосрентген 32": {"chat_id": -5147774326, "systems": ["Ремонт крыши"]},
    "Хутор Ильичевка 8": {"chat_id": -5256023040, "systems": ["Ремонт подъездов"]},
    "Троицк Центральная 4": {"chat_id": -5074236078, "systems": ["Ремонт подъездов"]},
    "Филимонковский Парковая 5": {"chat_id": -5044018898, "systems": ["Ремонт крыши", "Ремонт подъездов"]},
    "Филимонковский Центральная 1": {
        "chat_id": -5283494406,
        "systems": ["ГВС маг", "ХВС маг", "ЦО маг", "КН маг", "Ремонт подвала"],
    },
    "Филимонковский Центральная 3": {"chat_id": -5082118687, "systems": ["КН маг"]},
    "Филимонковский Горчаково Школьная 4": {
        "chat_id": -5010528170,
        "systems": ["Ремонт крыши", "Ремонт подъездов"],
    },
    "Филимонковский 1й Мкр 1": {"chat_id": -5220977732, "systems": ["Ремонт фасада"]},
    "Краснопахорский Шишкин Лес 23": {"chat_id": -5223853098, "systems": ["Ремонт фасада", "Ремонт Крыши"]},
    "Рогово Школьная 20": {
        "chat_id": -5110229686,
        "systems": [
            "ГВС маг",
            "ХВС маг",
            "ЦО маг",
            "КН маг",
            "ГВС ст",
            "ХВС ст",
            "ЦО ст",
            "КН st",
            "Ремонт подвала",
        ],
    },
    "Рогово Юбилейная 16": {
        "chat_id": -5218573114,
        "systems": [
            "ГВС маг",
            "ХВС маг",
            "ЦО маг",
            "КН маг",
            "ГВС ст",
            "ХВС ст",
            "ЦО ст",
            "КН ст",
            "Ремонт подвала",
            "Ремонт крыши",
            "Ремонт фасада",
        ],
    },
    "Кленово ул. Мичурина 2": {"chat_id": -5187871853, "systems": ["ХВС ст", "ГВС ст", "ЦО ст"]},
    "Щербинка Ерино 3": {"chat_id": -4993247238, "systems": ["ХВС маг", "ГВС маг", "ЦО мак", "КН маг", "Ремонт подвала"]},
    "Щербинка Знамя Октября 24": {
        "chat_id": -5108632810,
        "systems": ["ХВС маг", "ГВС маг", "ЦО маг", "КН маг", "Ремонт подвала", "Ремонт крыши"],
    },
    "Щербинка Фабрика им. 1 Мая 46": {"chat_id": -4964265480, "systems": ["Ремонт крыши"]},
    "Рекламация 2025": {"chat_id": -5044901573, "systems": ["Фото исправлений"]},
}

REKLAMACIA_CHAT_ID = -5044901573

COMMON_DOCS_BUTTON = "📁 Общие документы"
COMMON_DOCS_FOLDER = "_PROJECT"

SYSTEM_PROMPT = """
Ты — главный инженер и технический эксперт по капитальному ремонту многоквартирных домов (МКД) в Москве.
Твоя специализация: стандарты ФКР Москвы, регламенты ГБУ "Жилищник", строительные нормы (СП, СНиП, ГОСТ, ТР ТС).

Твои задачи:
1. Анализировать фото и вопросы на предмет нарушений технологий ФКР Москвы.
2. Ссылаться на конкретные пункты СП, Технических регламентов ФКР Москвы и альбомы технических решений.
3. Давать четкие рекомендации по устранению замечаний (как для технадзора).
4. При запросе схем или узлов — генерировать детальные текстовые описания чертежей или ASCII-схемы с указанием всех допустимых размеров, слоев и материалов (в мм).

Стиль общения: профессиональный, строгий, краткий, по существу.
Если видишь нарушение — пиши: "НАРУШЕНИЕ: <суть>. Ссылка на норму: <документ>. Как исправить: <инструкция>".
""".strip()


# -------------------- STATE (in-memory) --------------------
pending_photos = {}
pending_progress = {}
pending_deadline_setup = {}


# -------------------- JSON HELPERS --------------------
def _atomic_write_json(path: str, data: dict) -> None:
    tmp = f"{path}.tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def _load_json(path: str, default):
    if not os.path.exists(path):
        return default
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"JSON read error {path}: {e}")
        return default


# -------------------- DEADLINES --------------------
def load_deadlines() -> dict:
    return _load_json(DEADLINES_FILE, {})


def save_system_deadline(chat_id: int, system_name: str, date_str: str) -> None:
    data = load_deadlines()
    chat_key = str(chat_id)
    if chat_key not in data:
        data[chat_key] = {}
    data[chat_key][system_name] = date_str
    _atomic_write_json(DEADLINES_FILE, data)


def get_deadlines_report(chat_id: int, show_all: bool = False) -> str:
    data = load_deadlines()
    chat_data = data.get(str(chat_id), {})

    if not chat_data:
        return "" if not show_all else "📅 Сроки сдачи работ еще не установлены."

    report_lines = []
    today = datetime.now()
    weekday = today.weekday()

    def _key(item):
        try:
            return datetime.strptime(item[1], "%d.%m.%Y")
        except Exception:
            return datetime.max

    sorted_systems = sorted(chat_data.items(), key=_key)

    for system, date_str in sorted_systems:
        try:
            deadline_date = datetime.strptime(date_str, "%d.%m.%Y")
        except ValueError:
            continue

        delta = deadline_date - today
        days = delta.days + 1

        if show_all:
            show_reminder = True
        elif days <= 30:
            show_reminder = True
        else:
            show_reminder = weekday in (0, 4)

        if not show_reminder:
            continue

        if days > 5:
            status = f"✅ {days} дн."
        elif 0 <= days <= 5:
            status = f"⚠️ <b>{days} дн.</b>"
        else:
            status = f"🔥 <b>Просрочка {abs(days)} дн!</b>"

        report_lines.append(f"— {system}: {status} (до {date_str})")

    if not report_lines:
        return "" if not show_all else "📅 Нет активных напоминаний на сегодня."

    header = "📋 <b>Полный график работ:</b>\n" if show_all else "\n\n⏰ <b>Сроки окончания работ:</b>\n"
    return header + "\n".join(report_lines)


# -------------------- PROGRESS STATE --------------------
def load_progress_state() -> dict:
    return _load_json(PROGRESS_STATE_FILE, {})


def save_progress_state(data: dict) -> None:
    _atomic_write_json(PROGRESS_STATE_FILE, data)


def get_prev_progress(chat_id: int, system_name: str):
    state = load_progress_state()
    chat_key = str(chat_id)
    rec = state.get(chat_key, {}).get(system_name)
    if not rec:
        return 0.0, None
    return float(rec.get("last_percent", 0)), rec.get("last_date")


# -------------------- UI --------------------
def is_admin_user(update: Update) -> bool:
    uid = update.effective_user.id if update.effective_user else None
    return uid in ADMIN_USER_IDS


def build_groups_keyboard(selected: set):
    rows = []
    sorted_groups = sorted(GROUPS_CONFIG.items())

    for name, cfg in sorted_groups:
        if name == "Рекламация 2025":
            continue
        cid = cfg["chat_id"]
        mark = "✅" if cid in selected else "⬜️"
        rows.append([InlineKeyboardButton(f"{mark} {name}", callback_data=f"bc_tgl:{cid}")])

    rows.append(
        [
            InlineKeyboardButton("✅ Все", callback_data="bc_all"),
            InlineKeyboardButton("🧹 Сброс", callback_data="bc_none"),
        ]
    )
    rows.append(
        [
            InlineKeyboardButton("🚀 ОТПРАВИТЬ", callback_data="bc_done"),
            InlineKeyboardButton("❌ Отмена", callback_data="bc_cancel"),
        ]
    )
    return InlineKeyboardMarkup(rows)


MAIN_MENU = ReplyKeyboardMarkup(
    [
        [KeyboardButton("📤 Рассылка"), KeyboardButton("📋 Сроки работ")],
        [KeyboardButton("🆔 Мой ID")],
    ],
    resize_keyboard=True,
)


# -------------------- FILE HELPERS --------------------
def _clean_name(name: str) -> str:
    return "".join([c if c.isalnum() or c in "._- " else "_" for c in name]).strip()


def save_file_to_system(local_path: str, address: str, system: str, filename: str) -> str:
    address_clean = _clean_name(address)
    system_clean = _clean_name(system)

    dest_dir = os.path.join(PROJECTS_DIR, address_clean, system_clean)
    os.makedirs(dest_dir, exist_ok=True)

    dest_path = os.path.join(dest_dir, filename)
    shutil.copy(local_path, dest_path)
    return dest_path


def build_progress_keyboard(min_value: int):
    step = 5
    values = [v for v in range(0, 101, step) if v >= int(min_value)]

    rows = []
    row = []
    for v in values:
        row.append(InlineKeyboardButton(f"{v}%", callback_data=f"prog:{v}"))
        if len(row) == 5:
            rows.append(row)
            row = []
    if row:
        rows.append(row)

    return InlineKeyboardMarkup(rows)


def create_or_update_progress_excel(address: str, date_str: str, data: dict):
    address_clean = _clean_name(address)
    path = os.path.join(DATA_DIR, address_clean)
    os.makedirs(path, exist_ok=True)

    xlsx = os.path.join(path, "Прогресс_работ.xlsx")
    if os.path.exists(xlsx):
        wb = load_workbook(xlsx)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Прогресс"
        ws.append(["Дата", "Система", "Процент выполнения, %"])

    for sys, val in data.items():
        ws.append([date_str, sys, val])

    wb.save(xlsx)


# -------------------- AI --------------------
async def get_gpt_response(text: str, context: str | None = None) -> str:
    if context:
        system_msg = (
            f"{SYSTEM_PROMPT}\n\n"
            f"ВАЖНО: У тебя есть доступ к проектной документации и сметам объекта.\n"
            f"Используй следующую информацию для ответа:\n{context}\n\n"
            f"Если в документах нет ответа, так и скажи, но попробуй ответить как эксперт."
        )
    else:
        system_msg = SYSTEM_PROMPT

    try:
        r = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": system_msg},
                {"role": "user", "content": text},
            ],
            temperature=0.5,
        )
        return r.choices[0].message.content or ""
    except Exception as e:
        return f"⚠️ Ошибка: {str(e)}"


async def get_vision_response(text: str, image_path: str) -> str:
    try:
        with open(image_path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("utf-8")

        r = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": text or "Анализ фото"},
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}},
                    ],
                },
            ],
        )
        return r.choices[0].message.content or ""
    except Exception as e:
        return f"⚠️ Ошибка: {str(e)}"


# -------------------- HANDLERS --------------------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.type == "private":
        await update.message.reply_text(
            "👋 Привет! Я бот-помощник.\nВыбери действие в меню внизу:",
            reply_markup=MAIN_MENU,
        )
    else:
        await update.message.reply_text("👷‍♂️ Бот работает в группе!", parse_mode="HTML")


async def get_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    await update.message.reply_text(f"🆔 ID этого чата: <code>{chat_id}</code>", parse_mode="HTML")


async def start_deadline_setup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    chat_title = update.effective_chat.title or "Личный чат"

    target_cfg = None
    for _, cfg in GROUPS_CONFIG.items():
        if cfg["chat_id"] == chat_id:
            target_cfg = cfg
            break
    if not target_cfg and chat_title in GROUPS_CONFIG:
        target_cfg = GROUPS_CONFIG[chat_title]

    if not target_cfg:
        await update.message.reply_text("❌ Чат не настроен.")
        return

    systems = target_cfg["systems"]
    kb = [[InlineKeyboardButton(s, callback_data=f"deadline_{i}")] for i, s in enumerate(systems)]
    pending_deadline_setup[chat_id] = {"systems": systems}
    await update.message.reply_text("📅 Выберите систему:", reply_markup=InlineKeyboardMarkup(kb))


async def handle_deadline_system(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    cid = update.effective_chat.id
    idx = int(q.data.split("_")[1])

    if cid not in pending_deadline_setup:
        await q.edit_message_text("❌ Ошибка. Начните заново.")
        return

    sys = pending_deadline_setup[cid]["systems"][idx]
    pending_deadline_setup[cid]["sel_sys"] = sys
    await q.edit_message_text(f"Система: <b>{sys}</b>\nВведите дату (ДД.ММ.ГГГГ):", parse_mode="HTML")


async def show_deadlines_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    report = get_deadlines_report(chat_id, show_all=True)
    await update.message.reply_text(report, parse_mode="HTML")


async def handle_progress_button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    cid = update.effective_chat.id
    if cid not in pending_progress:
        await q.edit_message_text("❌ Опрос прогресса не активен. Дождитесь следующего запуска.")
        return

    try:
        new_val = float(q.data.split(":")[1])
    except Exception:
        await q.answer("❌ Ошибка данных кнопки.", show_alert=True)
        return

    st = pending_progress[cid]
    system = st["curr"]

    prev_val, prev_date = get_prev_progress(cid, system)
    if new_val < prev_val:
        await q.answer(f"Нельзя меньше чем было: {prev_val}% (дата {prev_date or '—'})", show_alert=True)
        return

    st["ans"][system] = new_val

    if st["left"]:
        nxt = st["left"].pop(0)
        st["curr"] = nxt

        pv, pd = get_prev_progress(cid, nxt)
        text = (
            f"🔧 Система: <b>{nxt}</b>\n"
            f"Предыдущее: <b>{pv}%</b> (дата: {pd or '—'})\n\n"
            f"👇 Выберите новый процент (не меньше предыдущего):"
        )
        await q.edit_message_text(text, parse_mode="HTML", reply_markup=build_progress_keyboard(pv))
        return

    title = update.effective_chat.title or "Личный чат"
    create_or_update_progress_excel(title, st["date"], st["ans"])

    state = load_progress_state()
    chat_key = str(cid)
    if chat_key not in state:
        state[chat_key] = {}

    for sys_name, val in st["ans"].items():
        state[chat_key][sys_name] = {"last_percent": float(val), "last_date": st["date"]}

    save_progress_state(state)

    del pending_progress[cid]
    await q.edit_message_text("✅ Прогресс сохранён. Спасибо!", parse_mode="HTML")


# -------------------- BROADCAST --------------------
async def broadcast_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin_user(update):
        await update.message.reply_text("⛔️ У вас нет прав на рассылку.")
        return

    context.user_data["bc_selected"] = set()
    context.user_data["bc_wait_message"] = False

    await update.message.reply_text(
        "📢 <b>Режим рассылки</b>\nОтметьте группы получателей:",
        reply_markup=build_groups_keyboard(context.user_data["bc_selected"]),
        parse_mode="HTML",
    )


async def broadcast_buttons(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if "bc_selected" not in context.user_data:
        context.user_data["bc_selected"] = set()

    selected = context.user_data["bc_selected"]

    if data.startswith("bc_tgl:"):
        chat_id = int(data.split(":")[1])
        if chat_id in selected:
            selected.remove(chat_id)
        else:
            selected.add(chat_id)
    elif data == "bc_all":
        for name, cfg in GROUPS_CONFIG.items():
            if name == "Рекламация 2025":
                continue
            selected.add(cfg["chat_id"])
    elif data == "bc_none":
        selected.clear()
    elif data == "bc_cancel":
        context.user_data.clear()
        await query.edit_message_text("❌ Рассылка отменена.")
        return
    elif data == "bc_done":
        if not selected:
            await query.answer("⚠️ Выберите хотя бы одну группу!", show_alert=True)
            return
        context.user_data["bc_wait_message"] = True
        group_count = len(selected)
        await query.edit_message_text(
            f"✅ Выбрано групп: <b>{group_count}</b>\n\n"
            f"✍️ <b>Теперь пришлите сообщение</b> (текст, фото или файл),\n"
            f"и я перешлю его в выбранные чаты.",
            parse_mode="HTML",
        )
        return

    try:
        await query.edit_message_reply_markup(reply_markup=build_groups_keyboard(selected))
    except Exception:
        pass


async def execute_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    targets = context.user_data.get("bc_selected", set())
    if not targets:
        await update.message.reply_text("❌ Ошибка: получатели не выбраны.")
        context.user_data.clear()
        return

    msg = update.message
    success = 0
    failed = 0
    status_msg = await update.message.reply_text("⏳ Рассылка запущена...")

    for chat_id in targets:
        try:
            if msg.text:
                await context.bot.send_message(chat_id=chat_id, text=msg.text)
            elif msg.photo:
                await context.bot.send_photo(chat_id=chat_id, photo=msg.photo[-1].file_id, caption=msg.caption)
            elif msg.video:
                await context.bot.send_video(chat_id=chat_id, video=msg.video.file_id, caption=msg.caption)
            elif msg.document:
                await context.bot.send_document(chat_id=chat_id, document=msg.document.file_id, caption=msg.caption)
            elif msg.audio:
                await context.bot.send_audio(chat_id=chat_id, audio=msg.audio.file_id, caption=msg.caption)
            elif msg.voice:
                await context.bot.send_voice(chat_id=chat_id, voice=msg.voice.file_id, caption=msg.caption)
            else:
                failed += 1
                continue

            success += 1
        except Exception as e:
            logger.error(f"Не удалось отправить в {chat_id}: {e}")
            failed += 1

    context.user_data.clear()
    await context.bot.edit_message_text(
        chat_id=update.effective_chat.id,
        message_id=status_msg.message_id,
        text=f"✅ <b>Рассылка завершена!</b>\n\nУспешно: {success}\nОшибок: {failed}",
        parse_mode="HTML",
    )


# -------------------- TEXT --------------------
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    cid = update.effective_chat.id
    title = update.effective_chat.title or "Личный чат"

    if text == "📤 Рассылка":
        await broadcast_start(update, context)
        return
    if text == "📋 Сроки работ":
        await show_deadlines_command(update, context)
        return
    if text == "🆔 Мой ID":
        await get_id(update, context)
        return

    if context.user_data.get("bc_wait_message"):
        await execute_broadcast(update, context)
        return

    if cid in pending_deadline_setup and "sel_sys" in pending_deadline_setup[cid]:
        sys = pending_deadline_setup[cid]["sel_sys"]
        try:
            datetime.strptime(text, "%d.%m.%Y")
            save_system_deadline(cid, sys, text)
            await update.message.reply_text(f"✅ Срок для <b>{sys}</b>: {text}", parse_mode="HTML")
            del pending_deadline_setup[cid]
        except ValueError:
            await update.message.reply_text("❌ Формат ДД.ММ.ГГГГ")
        return

    if text.startswith("*"):
        user_query = text[1:].strip()
        await update.message.chat.send_action("typing")

        project_name = None
        for p_name, cfg in GROUPS_CONFIG.items():
            if cfg["chat_id"] == cid:
                project_name = p_name
                break
        if not project_name and title in GROUPS_CONFIG:
            project_name = title

        context_data = None
        if project_name:
            context_data = rag_engine.get_relevant_context(project_name, user_query)

        res = await get_gpt_response(user_query, context=context_data)

        await context.bot.send_message(
            chat_id=cid,
            text=res or "⚠️ Не удалось получить ответ",
            parse_mode=None,
            disable_web_page_preview=True,
        )
        return

    if cid in pending_progress:
        await update.message.reply_text("📊 Для ввода прогресса используйте кнопки под сообщением.")
        return


# -------------------- DOCS REINDEX --------------------
async def reload_docs_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin_user(update):
        return

    msg = await update.message.reply_text("⏳ Начинаю переиндексацию всех проектов...")
    count = 0

    for project_name in GROUPS_CONFIG.keys():
        if rag_engine.build_index_for_project(project_name):
            count += 1

    await context.bot.edit_message_text(
        chat_id=update.effective_chat.id,
        message_id=msg.message_id,
        text=f"✅ База знаний обновлена! Проиндексировано проектов: {count}",
    )


# -------------------- MEDIA --------------------
async def handle_media(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.user_data.get("bc_wait_message"):
        await execute_broadcast(update, context)
        return

    msg = update.message
    chat_title = msg.chat.title or f"chat_{msg.chat.id}"
    chat_id = msg.chat.id
    message_id = msg.message_id

    caption = msg.caption or ""

    # Vision: caption начинается с "*"
    if caption.strip().startswith("*") and msg.photo:
        file_obj = await context.bot.get_file(msg.photo[-1].file_id)
        with tempfile.TemporaryDirectory() as temp_dir:
            local_path = os.path.join(temp_dir, f"{message_id}.jpg")
            await file_obj.download_to_drive(local_path)
            await msg.chat.send_action("typing")
            ai_answer = await get_vision_response(caption.replace("*", "").strip(), local_path)
            await msg.reply_text(ai_answer)
        return

    # Сохранение фото/файла в DATA_DIR
    target_cfg = None
    for _, cfg in GROUPS_CONFIG.items():
        if cfg["chat_id"] == chat_id:
            target_cfg = cfg
            break
    if not target_cfg and chat_title in GROUPS_CONFIG:
        target_cfg = GROUPS_CONFIG[chat_title]

    if not target_cfg:
        return

    file_obj = None
    file_ext = ""

    if msg.photo:
        file_obj = await context.bot.get_file(msg.photo[-1].file_id)
        file_ext = ".jpg"
    elif msg.document:
        file_obj = await context.bot.get_file(msg.document.file_id)
        if msg.document.file_name:
            _, ext = os.path.splitext(msg.document.file_name)
            file_ext = ext or ".bin"
        else:
            file_ext = ".bin"
    else:
        return

    temp_dir = tempfile.gettempdir()
    filename = f"stroybot_{chat_id}_{message_id}{file_ext}"
    local_path = os.path.join(temp_dir, filename)

    await file_obj.download_to_drive(local_path)

    pending_photos[f"{chat_id}_{message_id}"] = {
        "local_path": local_path,
        "filename": filename,
        "chat_title": chat_title,
        "config": target_cfg,
    }

    # добавляем общую папку проекта как первый пункт
    systems = [COMMON_DOCS_BUTTON] + target_cfg["systems"]
    keyboard = [[InlineKeyboardButton(s, callback_data=f"save_{chat_id}_{message_id}_{i}")] for i, s in enumerate(systems)]
    await msg.reply_text("🔧 К какой папке сохранить файл?", reply_markup=InlineKeyboardMarkup(keyboard))


async def handle_save_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    data = query.data.split("_")
    chat_id, message_id, sys_idx = int(data[1]), int(data[2]), int(data[3])
    key = f"{chat_id}_{message_id}"

    if key not in pending_photos:
        await query.edit_message_text("❌ Файл устарел или уже сохранён.")
        return

    d = pending_photos[key]

    systems = [COMMON_DOCS_BUTTON] + d["config"]["systems"]
    chosen = systems[sys_idx]
    folder_name = COMMON_DOCS_FOLDER if chosen == COMMON_DOCS_BUTTON else chosen

    dest_path = save_file_to_system(d["local_path"], d["chat_title"], folder_name, d["filename"])

    await query.edit_message_text(
        f"✅ Файл сохранён:\n<b>{chosen}</b>\n<code>{dest_path}</code>",
        parse_mode="HTML",
    )

    if os.path.exists(d["local_path"]):
        try:
            os.remove(d["local_path"])
        except Exception as e:
            logger.error(f"Не удалось удалить временный файл: {e}")

    del pending_photos[key]


# -------------------- JOBS --------------------
async def ask_for_system_progress(context: ContextTypes.DEFAULT_TYPE):
    d_str = datetime.now().strftime("%Y-%m-%d")

    for addr, cfg in GROUPS_CONFIG.items():
        cid = cfg.get("chat_id")
        if not cid:
            continue
        if cid == REKLAMACIA_CHAT_ID:
            continue

        deadlines_text = get_deadlines_report(cid)

        systems = cfg["systems"].copy()
        if not systems:
            continue

        first_sys = systems.pop(0)
        pending_progress[cid] = {"date": d_str, "left": systems, "ans": {}, "curr": first_sys}

        prev_val, prev_date = get_prev_progress(cid, first_sys)

        text = (
            f"📅 <b>Отчёт: {addr}</b>{deadlines_text}\n\n"
            f"🔧 Система: <b>{first_sys}</b>\n"
            f"Предыдущее: <b>{prev_val}%</b> (дата: {prev_date or '—'})\n\n"
            f"👇 Выберите новый процент (не меньше предыдущего):"
        )

        await context.bot.send_message(
            cid,
            text,
            parse_mode="HTML",
            reply_markup=build_progress_keyboard(prev_val),
        )


async def test_progress(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await ask_for_system_progress(context)


async def progress_report_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cid = update.effective_chat.id

    target_cfg = None
    for _, cfg in GROUPS_CONFIG.items():
        if cfg["chat_id"] == cid:
            target_cfg = cfg
            break
    if not target_cfg:
        await update.message.reply_text("❌ Чат не настроен в GROUPS_CONFIG.")
        return

    systems = target_cfg["systems"]
    state = load_progress_state().get(str(cid), {})

    lines = ["📊 <b>Свод прогресса:</b>"]
    for s in systems:
        rec = state.get(s)
        if rec:
            lines.append(f"— {s}: <b>{rec.get('last_percent', 0)}%</b> (дата: {rec.get('last_date', '—')})")
        else:
            lines.append(f"— {s}: <b>—</b> (дата: —)")

    await update.message.reply_text("\n".join(lines), parse_mode="HTML")


def _setup_jobs(app):
    msk_tz = pytz.timezone("Europe/Moscow")
    trigger = CronTrigger(day_of_week="tue,fri", hour=15, minute=0, second=0, timezone=msk_tz)

    app.job_queue.run_custom(
        ask_for_system_progress,
        job_kwargs={"trigger": trigger},
        name="progress_tue_fri",
    )


def main():
    logger.info("🚀 БОТ ЗАПУЩЕН...")
    app = ApplicationBuilder().token(TELEGRAM_TOKEN).build()

    rag_engine.configure(data_dir=DATA_DIR)

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("test_progress", test_progress))
    app.add_handler(CommandHandler("progress", progress_report_command))
    app.add_handler(CommandHandler("set_deadline", start_deadline_setup))
    app.add_handler(CommandHandler("deadlines", show_deadlines_command))
    app.add_handler(CommandHandler("get_id", get_id))
    app.add_handler(CommandHandler("broadcast", broadcast_start))
    app.add_handler(CommandHandler("reload_docs", reload_docs_command))

    app.add_handler(CallbackQueryHandler(handle_deadline_system, pattern="^deadline_"))
    app.add_handler(CallbackQueryHandler(handle_save_selection, pattern="^save_"))
    app.add_handler(CallbackQueryHandler(broadcast_buttons, pattern="^bc_"))
    app.add_handler(CallbackQueryHandler(handle_progress_button, pattern="^prog:"))

    app.add_handler(MessageHandler(filters.PHOTO | filters.VIDEO | filters.Document.ALL, handle_media))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    _setup_jobs(app)
    app.run_polling()


if __name__ == "__main__":
    main()



